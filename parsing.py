import openpyxl
from bs4 import BeautifulSoup
import pickle
import sys

doc = openpyxl.load_workbook('/home/nicolas/Documents/1001pact/TEST B1.xlsx')
print (type(doc))

sheet = doc.get_sheet_by_name('Nodes')
print(type(sheet))

mylist = []
for row in sheet.iter_rows('C{}:C{}'.format(sheet.min_row,sheet.max_row)):
    for cell in row:
        mylist.append(cell.value)


def create():
    print("creating new  file")
    name=input("enter the name of file:")
    extension=input("enter extension of file:")
    name=name+"."+extension
    file = open(name,"w")
    pickle.dump(mylist,open(name, "wb"))
    file.close()

create()

soup = BeautifulSoup(open("/home/nicolas/Documents/1001pact/Mapping_media/linklist.html", 'rb'), "lxml")
links = []
for link in soup.find_all('a', text='Website'):
    links.append(link.get('href'))

testSheet = doc.get_sheet_by_name('test')
print(type(testSheet))

r = 1
for i in links: 
    testSheet.cell(row=r, column=1).value = i
    r += 1
    
    

doc.save('Test B1.xlsx')



















